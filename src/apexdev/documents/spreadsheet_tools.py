"""Spreadsheet conversion utilities using openpyxl."""

from __future__ import annotations

from typing import Iterable, Any
from pathlib import Path
import csv

from ..core.models import Conversation

try:
    from openpyxl import Workbook, load_workbook  # type: ignore
except ImportError as exc:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore
    _xlsx_import_error = exc
else:
    _xlsx_import_error = None


def _require_openpyxl():
    if Workbook is None or load_workbook is None:
        raise ImportError("openpyxl is required for spreadsheet conversion") from _xlsx_import_error


def conversations_to_workbook(conversations: Iterable[Conversation], output_path: str | Path) -> None:
    if Workbook is None:
        raise ImportError("openpyxl is required for spreadsheet conversion but is not installed") from _xlsx_import_error
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    for convo in conversations:
        sheet_name = (convo.title or 'Conversation')[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(['Index', 'Role', 'Content'])
        for idx, msg in enumerate(convo.messages):
            ws.append([idx, msg.role, msg.content])
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def csv_to_xlsx(input_path: str | Path, output_path: str | Path, *, sheet_name: str = "Sheet1") -> Path:
    _require_openpyxl()
    src = Path(input_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    with src.open(newline="", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(str(out))
    return out


def xlsx_to_csv(input_path: str | Path, output_path: str | Path, *, sheet: str | int | None = None) -> Path:
    _require_openpyxl()
    src = Path(input_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(str(src), data_only=False, read_only=True)
    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    elif isinstance(sheet, str):
        ws = wb[sheet]
    else:
        ws = wb.worksheets[0]
    with out.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    return out


def xlsx_to_csvs(input_path: str | Path, output_dir: str | Path) -> list[Path]:
    _require_openpyxl()
    src = Path(input_path)
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(str(src), data_only=False, read_only=True)
    outputs: list[Path] = []
    for ws in wb.worksheets:
        safe = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in ws.title).strip("_") or "sheet"
        dst = out_dir / f"{safe}.csv"
        with dst.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if v is None else v for v in row])
        outputs.append(dst)
    return outputs
